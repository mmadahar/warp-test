#!/usr/bin/env python3
"""
Excel Workbook Verification Script

This script verifies the structure and data of an Excel workbook created
by the excel_generator.py script.
"""

import argparse
import os
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def verify_workbook(workbook_path: str) -> None:
    """
    Verify the structure and data of an Excel workbook.
    
    Args:
        workbook_path: Path to the Excel workbook to verify
    """
    print(f"\nVerifying workbook: {workbook_path}")
    print("=" * 80)
    
    # Check if the file exists
    if not os.path.exists(workbook_path):
        print(f"Error: File {workbook_path} does not exist.")
        return
    
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(workbook_path)
        
        # Print workbook information
        print(f"Workbook loaded successfully: {os.path.basename(workbook_path)}")
        print(f"Number of worksheets: {len(wb.worksheets)}")
        
        # Print worksheet names
        print("\nWorksheet names:")
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            print(f"  {i}. {sheet_name}")
        
        # Verify each worksheet has data
        print("\nWorksheet data verification:")
        print("-" * 60)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Get used range dimensions
            min_row, min_col = 1, 1
            max_row = ws.max_row
            max_col = ws.max_column
            
            # Count non-empty cells
            non_empty_count = 0
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        non_empty_count += 1
            
            print(f"\nSheet: {sheet_name}")
            print(f"  Dimensions: {max_row} rows x {max_col} columns")
            print(f"  Non-empty cells: {non_empty_count}")
            
            # Sample data - print first row (headers)
            if max_row > 0 and max_col > 0:
                print("  Headers:")
                for j in range(1, min(max_col + 1, 6)):  # First 5 columns or less
                    cell_value = ws.cell(row=1, column=j).value
                    print(f"    Column {j}: {cell_value}")
                
                # Print first few data rows
                if max_row > 1:
                    print("  Data samples (first 3 rows):")
                    for i in range(2, min(max_row + 1, 5)):  # Rows 2-4 (or less)
                        row_values = []
                        for j in range(1, min(max_col + 1, 4)):  # First 3 columns or less
                            cell_value = ws.cell(row=i, column=j).value
                            row_values.append(f"{cell_value}")
                        print(f"    Row {i}: {' | '.join(row_values)}")
        
        # Close the workbook
        wb.close()
        print("\nVerification complete.")
        
    except Exception as e:
        print(f"Error verifying workbook: {str(e)}")


def main():
    """Main function to parse arguments and verify workbook."""
    parser = argparse.ArgumentParser(
        description="Verify the structure and data of an Excel workbook."
    )
    parser.add_argument(
        "workbook_path",
        type=str,
        help="Path to the Excel workbook to verify"
    )
    
    args = parser.parse_args()
    verify_workbook(args.workbook_path)


if __name__ == "__main__":
    main()

