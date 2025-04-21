#!/usr/bin/env python3
"""
Excel Workbook Generator

This script generates a specified number of Excel workbooks with random worksheet names
and data. Each workbook contains exactly 10 worksheets.
"""

import argparse
import os
import random
import string
from pathlib import Path
from typing import List, Optional, Tuple

import openpyxl
from faker import Faker
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Import utility function
from src.utils import create_directory_if_not_exists


def generate_random_worksheet_name(faker: Faker, existing_names: List[str]) -> str:
    """
    Generate a random worksheet name that doesn't exist in the list of existing names.
    
    Args:
        faker: Faker instance for generating random data
        existing_names: List of existing worksheet names to avoid duplicates
    
    Returns:
        A unique random worksheet name
    """
    while True:
        # Generate a random name using various Faker methods
        name_type = random.randint(1, 5)
        if name_type == 1:
            name = faker.word().capitalize()
        elif name_type == 2:
            name = faker.job()
        elif name_type == 3:
            name = faker.city()
        elif name_type == 4:
            name = faker.country()
        else:
            name = faker.company()[:20]  # Limit length
            
        # Clean up the name to be a valid Excel worksheet name
        # Replace invalid characters with underscores
        name = ''.join(c if c.isalnum() or c in ' _' else '_' for c in name)
        name = name.strip()[:31]  # Excel worksheet names are limited to 31 chars
        
        # Ensure the name is not empty and not already used
        if name and name not in existing_names:
            existing_names.append(name)
            return name


def generate_random_data(worksheet: Worksheet, faker: Faker) -> None:
    """
    Populate a worksheet with random data.
    
    Args:
        worksheet: The worksheet to populate
        faker: Faker instance for generating random data
    """
    # Generate random number of rows and columns (between 5 and 20)
    rows = random.randint(5, 20)
    cols = random.randint(5, 10)
    
    # Add headers
    headers = []
    for i in range(cols):
        header_type = random.randint(1, 4)
        if header_type == 1:
            headers.append(faker.word().capitalize())
        elif header_type == 2:
            headers.append(faker.currency_name())
        elif header_type == 3:
            headers.append(faker.job().split()[0])
        else:
            headers.append(f"Column {string.ascii_uppercase[i]}")
    
    for col_idx, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_idx, value=header)
    
    # Add data rows
    for row_idx in range(2, rows + 2):
        for col_idx in range(1, cols + 1):
            data_type = random.randint(1, 5)
            if data_type == 1:
                value = faker.name()
            elif data_type == 2:
                value = random.randint(1, 1000)
            elif data_type == 3:
                value = faker.date()
            elif data_type == 4:
                value = faker.company()
            else:
                value = faker.sentence(nb_words=5)
            
            worksheet.cell(row=row_idx, column=col_idx, value=value)


def generate_workbooks(write_folder: str, num_workbooks: int = 1000) -> Tuple[bool, Optional[str]]:
    """
    Generate Excel workbooks with random worksheet names and data.
    
    Args:
        write_folder: Directory where workbooks will be saved
        num_workbooks: Number of workbooks to generate (default: 1000)
    
    Returns:
        A tuple (success, error_message) where success is a boolean indicating
        if all workbooks were created successfully, and error_message is a
        string with an error message if applicable.
    """
    # Create the directory if it doesn't exist
    try:
        create_directory_if_not_exists(write_folder)
    except Exception as e:
        return False, f"Failed to create directory: {str(e)}"
    
    # Set up the Faker instance for consistent random data
    faker = Faker()
    Faker.seed(random.randint(1, 10000))  # Set a random seed
    
    # Count of successfully created workbooks
    successful_count = 0
    
    try:
        for i in range(1, num_workbooks + 1):
            # Create a new workbook
            wb = Workbook()
            
            # Remove the default worksheet
            default_sheet = wb.active
            wb.remove(default_sheet)
            
            # Create a list to track worksheet names to avoid duplicates
            worksheet_names = []
            
            # Create 10 worksheets with random names and data
            for j in range(10):
                sheet_name = generate_random_worksheet_name(faker, worksheet_names)
                ws = wb.create_sheet(title=sheet_name)
                generate_random_data(ws, faker)
            
            # Save the workbook
            file_path = os.path.join(write_folder, f"workbook_{i:04d}.xlsx")
            wb.save(file_path)
            
            # Close workbook to free resources
            wb.close()
            
            successful_count += 1
            
            # Print progress every 100 workbooks
            if i % 100 == 0 or i == num_workbooks:
                print(f"Generated {i} of {num_workbooks} workbooks")
    
    except Exception as e:
        return False, f"Failed after creating {successful_count} workbooks: {str(e)}"
    
    return True, None


def main():
    """Main function to parse arguments and generate workbooks."""
    parser = argparse.ArgumentParser(
        description="Generate Excel workbooks with random worksheet names and data."
    )
    parser.add_argument(
        "--write_folder", 
        type=str, 
        required=True,
        help="Directory where workbooks will be saved"
    )
    parser.add_argument(
        "--num_workbooks", 
        type=int, 
        default=1000,
        help="Number of workbooks to generate (default: 1000)"
    )
    
    args = parser.parse_args()
    
    print(f"Generating {args.num_workbooks} workbooks in '{args.write_folder}'...")
    success, error_message = generate_workbooks(args.write_folder, args.num_workbooks)
    
    if success:
        print(f"Successfully generated {args.num_workbooks} workbooks in '{args.write_folder}'")
    else:
        print(f"Error: {error_message}")
        return 1
    
    return 0


if __name__ == "__main__":
    exit(main())

